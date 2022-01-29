'''
Author: ywyz
Date: 2022-01-25 09:52:01
LastEditTime: 2022-01-29 11:31:00
LastEditors: ywyz
Description: 将数据导出至mysql数据库
FilePath: /StudentsInfo/ImportDataToSQL.py
'''

from openpyxl import load_workbook
from Students import Students
import docx
from docx import Document


def InsertOldInfo():
    # 插入旧表
    filepath = '2020.9.1防疫信息表.xlsx'
    wb = load_workbook(filename=filepath)
    sheetName = "学生"
    ws = wb[sheetName]

    for row in ws.rows:
        if row[0].value == "序号":
            continue

        else:

            student = Students(row[3].value, '男', row[7].value, row[8].value,
                               row[4].value, row[5].value, row[11].value, ' ',
                               '0')
            student.InsertIntoSql()
    print("插入成功")


def insertNewInfo():
    """插入新表"""
    filepath = '防疫信息填报表新生.xlsx'
    wb = load_workbook(filename=filepath)
    sheetName = "Sheet1"
    ws = wb[sheetName]

    for row in ws.rows:
        student = Students(row[2].value, '男', '小班', ' ', row[3].value,
                           row[4].value, row[7].value, ' ', '0')
        student.InsertIntoSql()
    print("插入成功")


def DataTOWord():
    '''
    导出幼儿信息排查表中数据
    '''

    path = '幼儿信息排查表20210506副本.docx'
    docus = Document(path)
    tables = docus.tables
    table = tables[0]
    print("读取表格成功")
    for i in range(0, len(table.rows)):  # 遍历行
        for n in range(0, 13):
            result = table.cell(i, n).text
            print(result, end=" ")
        print()
